from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import os
import tempfile
from datetime import datetime
from modules.assistant import DoubanAssistant

app = Flask(__name__)

# 加载数据
# 尝试多个可能的文件路径
possible_paths = [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'AMIS-2025.csv'),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'AMIS-2025.csv')
]

csv_path = None
for path in possible_paths:
    if os.path.exists(path):
        csv_path = path
        break

if csv_path is None:
    raise FileNotFoundError(f"找不到AMIS-2025.csv文件，请确保文件位于以下路径之一: {possible_paths}")

df = pd.read_csv(csv_path)

# 重命名列以便于处理
df.rename(columns={
    'Source': 'source',
    'Country/Region': 'country',
    'Commodity': 'commodity',
    'Element': 'element',
    'Season': 'year',
    'Value': 'value',
    'Element Unit': 'unit'
}, inplace=True)

# 处理年份数据，从"2000/01"格式转换为"2000"
df['year'] = df['year'].str.split('/').str[0].astype(int)


# 添加数据验证函数
def validate_data_combination(country, commodity, element):
    """验证国家、商品和指标的组合是否在数据中存在"""
    if not country or not commodity or not element:
        return False
    
    # 如果是多个国家，检查每个国家
    if isinstance(country, list):
        for c in country:
            # 检查该国家是否有该商品
            country_data = df[df['country'] == c]
            if commodity not in country_data['commodity'].unique():
                return False
            
            # 检查该国家的该商品是否有该指标
            country_commodity_data = country_data[country_data['commodity'] == commodity]
            if element not in country_commodity_data['element'].unique():
                return False
        return True
    else:
        # 单个国家的情况
        country_data = df[df['country'] == country]
        if commodity not in country_data['commodity'].unique():
            return False
        
        country_commodity_data = country_data[country_data['commodity'] == commodity]
        return element in country_commodity_data['element'].unique()

# 创建数据组合查找表，提高查询效率
valid_combinations = {}
for country in df['country'].unique():
    valid_combinations[country] = {}
    country_data = df[df['country'] == country]
    
    for commodity in country_data['commodity'].unique():
        valid_combinations[country][commodity] = []
        commodity_data = country_data[country_data['commodity'] == commodity]
        
        for element in commodity_data['element'].unique():
            valid_combinations[country][commodity].append(element)

print(f"已创建有效数据组合查找表")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/analysis')
def analysis():
    return render_template('analysis.html')

@app.route('/map')
def map():
    return render_template('map.html')

@app.route('/explore')
def explore_page():
    """数据探索页面"""
    return render_template('explore.html')

@app.route('/assistant')
def assistant():
    return render_template('assistant.html')

# API端点
@app.route('/api/sample-data/', methods=['GET'])
def get_sample_data():
    """获取数据样本和基本统计信息"""
    # 获取样本数据（最多100条）
    sample = df.sample(min(100, len(df))).copy()
    
    # 转换年份为字符串类型，以便JSON序列化
    if 'year' in sample.columns:
        sample['year'] = sample['year'].astype(str)
    
    # 计算基本统计信息
    stats = {
        'total_rows': len(df),
        'countries_count': len(df['country'].unique()),
        'commodities_count': len(df['commodity'].unique()),
        'elements_count': len(df['element'].unique()),
        'elements': sorted(df['element'].unique().tolist()),
        'commodities': sorted(df['commodity'].unique().tolist()),
        'years_range': [int(df['year'].min()), int(df['year'].max())],
        'top_countries': df['country'].value_counts().head(10).to_dict()
    }
    
    return jsonify({
        'sample': sample.to_dict(orient='records'),
        'stats': stats
    })

@app.route('/api/data-quality/', methods=['GET'])
def get_data_quality():
    """获取数据质量信息"""
    # 计算缺失值
    missing_values = df.isnull().sum().to_dict()
    
    # 计算各指标的数据范围
    element_ranges = {}
    for element in df['element'].unique():
        element_data = df[df['element'] == element]
        if len(element_data) > 0:
            # 获取单位
            unit = element_data['unit'].iloc[0] if 'unit' in element_data.columns else ""
            
            # 计算统计值
            element_ranges[element] = {
                'min': float(element_data['value'].min()),
                'max': float(element_data['value'].max()),
                'mean': float(element_data['value'].mean()),
                'median': float(element_data['value'].median()),
                'unit': unit
            }
    
    # 计算每个商品的可用指标
    commodity_elements = {}
    for commodity in df['commodity'].unique():
        commodity_data = df[df['commodity'] == commodity]
        commodity_elements[commodity] = sorted(commodity_data['element'].unique().tolist())
    
    # 计算每个国家的可用商品
    country_commodities = {}
    for country in df['country'].unique():
        country_data = df[df['country'] == country]
        country_commodities[country] = sorted(country_data['commodity'].unique().tolist())
    
    # 计算每个国家和商品组合的可用指标
    country_commodity_elements = {}
    for country in df['country'].unique():
        country_commodity_elements[country] = {}
        country_data = df[df['country'] == country]
        for commodity in country_data['commodity'].unique():
            commodity_data = country_data[country_data['commodity'] == commodity]
            country_commodity_elements[country][commodity] = sorted(commodity_data['element'].unique().tolist())
    
    return jsonify({
        'missing_values': missing_values,
        'element_ranges': element_ranges,
        'commodity_elements': commodity_elements,
        'country_commodities': country_commodities,
        'country_commodity_elements': country_commodity_elements
    })

@app.route('/api/element-distribution/', methods=['GET'])
def get_element_distribution():
    """获取指标分布信息"""
    element = request.args.get('element', '')
    
    if not element:
        return jsonify({'error': '缺少指标参数'})
    
    # 筛选指定指标的数据
    element_data = df[df['element'] == element].copy()
    
    if len(element_data) == 0:
        return jsonify({'error': '未找到指定指标的数据'})
    
    # 获取单位
    unit = element_data['unit'].iloc[0] if 'unit' in element_data.columns else ""
    
    # 按商品分组计算平均值
    commodity_distribution = element_data.groupby('commodity')['value'].mean().reset_index()
    commodity_distribution = commodity_distribution.sort_values('value', ascending=False)
    
    # 按国家分组计算平均值
    country_distribution = element_data.groupby('country')['value'].mean().reset_index()
    country_distribution = country_distribution.sort_values('value', ascending=False)
    
    # 按年份分组计算平均值
    year_trend = element_data.groupby('year')['value'].mean().reset_index()
    year_trend = year_trend.sort_values('year')
    
    return jsonify({
        'element': element,
        'unit': unit,
        'total_records': len(element_data),
        'commodity_distribution': {
            'labels': commodity_distribution['commodity'].tolist(),
            'values': commodity_distribution['value'].tolist()
        },
        'country_distribution': {
            'labels': country_distribution.head(20)['country'].tolist(),  # 只取前20个国家
            'values': country_distribution.head(20)['value'].tolist()
        },
        'year_trend': {
            'years': year_trend['year'].astype(str).tolist(),
            'values': year_trend['value'].tolist()
        }
    })

@app.route('/api/metadata/', methods=['GET'])
def get_metadata():
    """获取数据元数据"""
    # 获取唯一值
    countries = sorted(df['country'].unique().tolist())
    commodities = sorted(df['commodity'].unique().tolist())
    elements = sorted(df['element'].unique().tolist())
    years = sorted(df['year'].unique().tolist())
    
    return jsonify({
        'countries': countries,
        'commodities': commodities,
        'elements': elements,
        'years': years,
        'valid_combinations': valid_combinations  # 添加有效组合信息
    })

@app.route('/api/analyze-data/', methods=['POST'])
def analyze_data():
    """分析数据API"""
    data = request.json
    
    # 获取请求参数
    countries = data.get('countries', [])
    commodity = data.get('commodity', '')
    element = data.get('element', '')
    start_year = data.get('start_year', '')
    end_year = data.get('end_year', '')
    
    # 验证参数
    if not countries or not commodity or not element:
        return jsonify({'error': '缺少必要参数'})
    
    # 验证数据组合是否有效
    valid = True
    invalid_countries = []
    for country in countries:
        if country not in valid_combinations or \
           commodity not in valid_combinations[country] or \
           element not in valid_combinations[country][commodity]:
            valid = False
            invalid_countries.append(country)
    
    if not valid:
        return jsonify({
            'error': f'无效的数据组合。以下国家没有 {commodity} 的 {element} 数据: {", ".join(invalid_countries)}',
            'data': [],
            'stats': {
                'total_records': 0,
                'countries_count': 0,
                'year_range': [0, 0],
                'by_country': {},
                'by_year': {}
            }
        })
    
    # 构建查询条件
    query = (df['country'].isin(countries)) & (df['commodity'] == commodity) & (df['element'] == element)
    
    # 添加年份范围条件
    if start_year:
        query &= (df['year'] >= int(start_year))
    if end_year:
        query &= (df['year'] <= int(end_year))
    
    # 筛选数据
    filtered_data = df[query].copy()
    
    if len(filtered_data) == 0:
        return jsonify({
            'data': [],
            'stats': {
                'total_records': 0,
                'countries_count': 0,
                'year_range': [0, 0],
                'by_country': {},
                'by_year': {}
            }
        })
    
    # 转换年份为字符串类型，以便JSON序列化
    filtered_data['year'] = filtered_data['year'].astype(str)
    
    # 准备返回数据
    result_data = filtered_data.to_dict(orient='records')
    
    # 计算统计信息
    stats = {
        'total_records': len(filtered_data),
        'countries_count': len(countries),
        'year_range': [
            filtered_data['year'].astype(int).min(),
            filtered_data['year'].astype(int).max()
        ],
        'by_country': {},
        'by_year': {}
    }
    
    # 按国家统计
    for country in countries:
        country_data = filtered_data[filtered_data['country'] == country]
        if len(country_data) > 0:
            stats['by_country'][country] = {
                'count': len(country_data),
                'mean': float(country_data['value'].mean()),
                'min': float(country_data['value'].min()),
                'max': float(country_data['value'].max())
            }
    
    # 按年份统计
    for year in sorted(filtered_data['year'].unique()):
        year_data = filtered_data[filtered_data['year'] == year]
        year_stats = {
            'count': len(year_data),
            'mean': float(year_data['value'].mean()),
            'min': float(year_data['value'].min()),
            'max': float(year_data['value'].max()),
            'by_country': {}
        }
        
        # 按国家和年份统计
        for country in countries:
            country_year_data = year_data[year_data['country'] == country]
            if len(country_year_data) > 0:
                year_stats['by_country'][country] = float(country_year_data['value'].mean())
        
        stats['by_year'][year] = year_stats
    
    return jsonify({
        'data': result_data,
        'stats': stats
    })

@app.route('/api/export-data/', methods=['POST'])
def export_data():
    """导出数据API"""
    data = request.json
    
    # 获取请求参数
    analysis_data = data.get('data', {})
    export_format = data.get('format', 'csv')
    include_stats = data.get('include_stats', True)
    include_chart = data.get('include_chart', True)
    
    if not analysis_data or not analysis_data.get('data'):
        return jsonify({'error': '没有可导出的数据'})
    
    # 创建临时文件
    import tempfile
    import os
    from datetime import datetime
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"agricultural_data_export_{timestamp}"
    
    # 根据导出格式处理
    if export_format == 'csv':
        # 导出为CSV
        import csv
        
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"{filename}.csv")
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['country', 'commodity', 'element', 'year', 'value', 'unit']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for row in analysis_data['data']:
                writer.writerow(row)
            
            # 添加统计信息
            if include_stats:
                writer.writerow({})
                writer.writerow({'country': '统计信息'})
                writer.writerow({'country': '总记录数', 'commodity': analysis_data['stats']['total_records']})
                writer.writerow({'country': '国家/地区数', 'commodity': analysis_data['stats']['countries_count']})
                writer.writerow({'country': '年份范围', 'commodity': f"{analysis_data['stats']['year_range'][0]} - {analysis_data['stats']['year_range'][1]}"})
                
                writer.writerow({})
                writer.writerow({'country': '按国家统计'})
                writer.writerow({'country': '国家/地区', 'commodity': '平均值', 'element': '最小值', 'year': '最大值'})
                
                for country, stats in analysis_data['stats']['by_country'].items():
                    writer.writerow({
                        'country': country,
                        'commodity': stats['mean'],
                        'element': stats['min'],
                        'year': stats['max']
                    })
        
        # 返回文件URL
        file_url = f"/download/{os.path.basename(file_path)}"
        
    elif export_format == 'excel':
        # 导出为Excel
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"{filename}.xlsx")
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"
        
        # 添加表头
        headers = ['国家/地区', '商品', '指标', '年份', '值', '单位']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # 添加数据
        for row_num, row in enumerate(analysis_data['data'], 2):
            ws.cell(row=row_num, column=1, value=row['country'])
            ws.cell(row=row_num, column=2, value=row['commodity'])
            ws.cell(row=row_num, column=3, value=row['element'])
            ws.cell(row=row_num, column=4, value=row['year'])
            ws.cell(row=row_num, column=5, value=row['value'])
            ws.cell(row=row_num, column=6, value=row['unit'])
        
        # 添加统计信息
        if include_stats:
            stats_ws = wb.create_sheet(title="统计信息")
            
            stats_ws.cell(row=1, column=1, value="基本统计")
            stats_ws.cell(row=2, column=1, value="总记录数")
            stats_ws.cell(row=2, column=2, value=analysis_data['stats']['total_records'])
            stats_ws.cell(row=3, column=1, value="国家/地区数")
            stats_ws.cell(row=3, column=2, value=analysis_data['stats']['countries_count'])
            stats_ws.cell(row=4, column=1, value="年份范围")
            stats_ws.cell(row=4, column=2, value=f"{analysis_data['stats']['year_range'][0]} - {analysis_data['stats']['year_range'][1]}")
            
            stats_ws.cell(row=6, column=1, value="按国家统计")
            stats_ws.cell(row=7, column=1, value="国家/地区")
            stats_ws.cell(row=7, column=2, value="平均值")
            stats_ws.cell(row=7, column=3, value="最小值")
            stats_ws.cell(row=7, column=4, value="最大值")
            
            row_num = 8
            for country, stats in analysis_data['stats']['by_country'].items():
                stats_ws.cell(row=row_num, column=1, value=country)
                stats_ws.cell(row=row_num, column=2, value=stats['mean'])
                stats_ws.cell(row=row_num, column=3, value=stats['min'])
                stats_ws.cell(row=row_num, column=4, value=stats['max'])
                row_num += 1
            
            # 添加按年份统计
            stats_ws.cell(row=row_num + 1, column=1, value="按年份统计")
            stats_ws.cell(row=row_num + 2, column=1, value="年份")
            
            # 添加国家列
            col_num = 2
            for country in analysis_data['stats']['by_country'].keys():
                stats_ws.cell(row=row_num + 2, column=col_num, value=country)
                col_num += 1
            
            stats_ws.cell(row=row_num + 2, column=col_num, value="平均值")
            
            # 添加年份数据
            row_num += 3
            for year, year_data in sorted(analysis_data['stats']['by_year'].items()):
                stats_ws.cell(row=row_num, column=1, value=year)
                
                col_num = 2
                for country in analysis_data['stats']['by_country'].keys():
                    value = year_data['by_country'].get(country, None)
                    stats_ws.cell(row=row_num, column=col_num, value=value)
                    col_num += 1
                
                stats_ws.cell(row=row_num, column=col_num, value=year_data['mean'])
                row_num += 1
        
        # 添加图表
        if include_chart and len(analysis_data['data']) > 0:
            chart_ws = wb.create_sheet(title="图表")
            
            # 准备图表数据
            chart_ws.cell(row=1, column=1, value="年份")
            
            # 添加国家列
            col_num = 2
            for country in analysis_data['stats']['by_country'].keys():
                chart_ws.cell(row=1, column=col_num, value=country)
                col_num += 1
            
            # 添加年份数据
            row_num = 2
            for year, year_data in sorted(analysis_data['stats']['by_year'].items()):
                chart_ws.cell(row=row_num, column=1, value=year)
                
                col_num = 2
                for country in analysis_data['stats']['by_country'].keys():
                    value = year_data['by_country'].get(country, None)
                    chart_ws.cell(row=row_num, column=col_num, value=value)
                    col_num += 1
                
                row_num += 1
            
            # 创建折线图
            chart = LineChart()
            chart.title = f"{analysis_data['data'][0]['commodity']} {analysis_data['data'][0]['element']}趋势比较"
            chart.style = 2
            chart.x_axis.title = "年份"
            chart.y_axis.title = analysis_data['data'][0]['unit']
            
            # 添加数据
            data = Reference(chart_ws, min_col=2, min_row=1, max_col=col_num-1, max_row=row_num-1)
            cats = Reference(chart_ws, min_col=1, min_row=2, max_row=row_num-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            chart_ws.add_chart(chart, "A10")
        
        # 保存Excel文件
        wb.save(file_path)
        
        # 返回文件URL
        file_url = f"/download/{os.path.basename(file_path)}"
        
    elif export_format == 'json':
        # 导出为JSON
        import json
        
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"{filename}.json")
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(analysis_data, jsonfile, ensure_ascii=False, indent=2)
        
        # 返回文件URL
        file_url = f"/download/{os.path.basename(file_path)}"
        
    else:
        return jsonify({'error': '不支持的导出格式'})
    
    # 添加下载路由
    @app.route(f'/download/{os.path.basename(file_path)}')
    def download_file():
        return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
    
    return jsonify({
        'success': True,
        'file_url': file_url,
        'filename': os.path.basename(file_path)
    })

@app.route('/api/map-data/', methods=['GET'])
def get_map_data():
    """获取地图数据"""
    # 获取请求参数
    commodity = request.args.get('commodity', '')
    element = request.args.get('element', '')
    year = request.args.get('year', '')
    
    # 验证参数
    if not commodity or not element or not year:
        return jsonify({'error': '缺少必要参数'})
    
    # 构建查询条件
    query = (df['commodity'] == commodity) & (df['element'] == element) & (df['year'] == int(year))
    
    # 筛选数据
    filtered_data = df[query].copy()
    
    if len(filtered_data) == 0:
        return jsonify({
            'error': '未找到符合条件的数据'
        })
    
    # 获取单位信息
    unit = filtered_data['unit'].iloc[0] if 'unit' in filtered_data.columns else ""
    
    # 转换年份为字符串类型，以便JSON序列化
    filtered_data['year'] = filtered_data['year'].astype(str)
    
    # 准备返回数据
    result_data = filtered_data.to_dict(orient='records')
    
    return jsonify({
        'data': result_data,
        'unit': unit
    })

@app.route('/api/chat/', methods=['POST'])
def chat():
    """与AI助手对话"""
    data = request.json
    user_message = data.get('message', '')
    
    assistant = DoubanAssistant()
    result = assistant.ask(user_message)
    
    if result.get('success'):
        # 生成数据洞察
        insights = assistant.generate_insights(user_message, result.get('reply'), df)
        
        return jsonify({
            'success': True,
            'reply': result.get('reply'),
            'insights': insights
        })
    else:
        return jsonify({
            'success': False,
            'error': result.get('error', '未知错误')
        })

@app.route('/api/data-summary/', methods=['GET'])
def get_data_summary():
    """获取数据摘要信息"""
    assistant = DoubanAssistant()
    summary = assistant.get_data_summary(df)
    
    if summary:
        return jsonify(summary)
    else:
        return jsonify({'error': '无法获取数据摘要'}), 500

if __name__ == '__main__':
    app.run(debug=True)


# 添加仪表盘数据API端点
@app.route('/api/dashboard-data/', methods=['GET'])
def get_dashboard_data():
    """获取仪表盘数据"""
    # 计算各年份数据量
    year_counts = df['year'].value_counts().sort_index()
    
    # 计算各商品数据量
    commodity_counts = df['commodity'].value_counts().head(10)
    
    # 计算各指标数据量
    element_counts = df['element'].value_counts().head(10)
    
    # 计算各国家/地区数据量
    country_counts = df['country'].value_counts().head(10)
    
    # 计算近5年主要商品趋势
    recent_years = sorted(df['year'].unique())[-5:]
    main_commodities = df['commodity'].value_counts().head(5).index.tolist()
    
    commodity_trends = {}
    for commodity in main_commodities:
        commodity_data = df[df['commodity'] == commodity]
        yearly_data = []
        
        for year in recent_years:
            year_data = commodity_data[commodity_data['year'] == year]
            if len(year_data) > 0:
                # 使用Production或Area Harvested指标，如果有的话
                for preferred_element in ['Production', 'Area Harvested', 'Yield']:
                    element_data = year_data[year_data['element'] == preferred_element]
                    if len(element_data) > 0:
                        yearly_data.append({
                            'year': str(year),
                            'value': float(element_data['value'].mean()),
                            'element': preferred_element,
                            'unit': element_data['unit'].iloc[0] if 'unit' in element_data.columns else ""
                        })
                        break
                else:
                    # 如果没有首选指标，使用任意指标
                    yearly_data.append({
                        'year': str(year),
                        'value': float(year_data['value'].mean()),
                        'element': year_data['element'].iloc[0],
                        'unit': year_data['unit'].iloc[0] if 'unit' in year_data.columns else ""
                    })
        
        if yearly_data:
            commodity_trends[commodity] = yearly_data
    
    return jsonify({
        'year_distribution': {
            'labels': year_counts.index.astype(str).tolist(),
            'values': year_counts.tolist()
        },
        'commodity_distribution': {
            'labels': commodity_counts.index.tolist(),
            'values': commodity_counts.tolist()
        },
        'element_distribution': {
            'labels': element_counts.index.tolist(),
            'values': element_counts.tolist()
        },
        'country_distribution': {
            'labels': country_counts.index.tolist(),
            'values': country_counts.tolist()
        },
        'commodity_trends': commodity_trends
    })